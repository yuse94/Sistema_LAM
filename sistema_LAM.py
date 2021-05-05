"""
LAM is under construction
Console Version 1.0 python3
"""

# LIBRARIES

from json import load, dump
from matplotlib.pyplot import show, figure, title, plot, savefig, close
from os import path, mkdir, system, getcwd
from shutil import copyfile
from time import strftime, sleep
from datetime import datetime, date
from pandas import DataFrame, ExcelWriter
from numpy import max, min, arctan, pi, arange, meshgrid, hstack, angle, array, sqrt, mean
from warnings import simplefilter
from webbrowser import open_new
from skimage import io, filters, morphology, transform, img_as_ubyte
from skimage.color import rgb2gray
from skimage.measure import label, regionprops
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

###############################
from tkinter import *
from tkinter import filedialog
##############################

from arrow import utcnow
from reportlab.platypus import PageBreak
from reportlab.platypus import Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.colors import black, whitesmoke, cornflowerblue, lightsteelblue
from reportlab.pdfgen import canvas
from reportlab.lib import utils

simplefilter(action='ignore', category=FutureWarning)


# FUNCTIONS TO IMAGE PROCESSING

def scale_image(img):
    """
    :param img:
    :return img_as_ubyte(img):
    """

    if len(img[1]) > 1000 or len(img[0]) > 2000:  # Maximum image size: X, Y
        img = transform.rescale(img, 0.8, multichannel=True)
        img = scale_image(img)
    return img_as_ubyte(img)


def green_filter(img, my_threshold):
    """
    :param img:
    :param my_threshold:
    :return filtered_image:
    """

    channel_green = rgb2gray(img[:, :, 1])
    gray_img = rgb2gray(img) * 255

    green_img = channel_green - gray_img

    green_img[green_img < 0] = 0

    green_img = filters.median(green_img)

    filter_threshold = round(max(green_img) * my_threshold)
    green_binary_img = green_img > filter_threshold
    filtered_image = morphology.remove_small_holes(green_binary_img)

    return filtered_image


def straighten_img(img, my_threshold):
    """
    :param img:
    :param my_threshold:
    :return straightened_img:
    """

    filtered_image = green_filter(img, my_threshold)

    img_markers = markers(filtered_image)
    reference_1, reference_2 = img_markers[0], img_markers[1]

    pos_reference = reference_2 - reference_1
    img_rotation_angle = arctan(pos_reference[0] / pos_reference[1])
    rotated_image = transform.rotate(img, img_rotation_angle * 180 / pi)

    reference_markers = [reference_1[1], reference_2[1]]
    reference_markers.sort()

    clipping_space_image = abs(reference_1[1] - reference_2[1]) / 10

    x_1 = int(round(reference_markers[0] - clipping_space_image))
    x_2 = int(round(reference_markers[1] + clipping_space_image))

    straightened_img = rotated_image[:, x_1:x_2]

    straightened_img = img_as_ubyte(straightened_img)

    return straightened_img


def img_grid(center_x, center_y, size_x, size_y, divisions_size):
    """
    :param center_x:
    :param center_y:
    :param size_x:
    :param size_y:
    :param divisions_size:
    :return mesh_x, mesh_y, x_horizontal, y_horizontal, x_vertical, y_vertical:
    """

    # The grid is built in the four quadrants of the image

    x_1 = arange(center_x, size_x, divisions_size)
    y_1 = arange(center_y, size_y, divisions_size)
    [mesh_x_1, mesh_y_1] = meshgrid(x_1, y_1)

    x_2 = arange(center_x, 0, -divisions_size)
    y_2 = arange(center_y, 0, -divisions_size)
    [mesh_x_2, mesh_y_2] = meshgrid(x_2, y_2)

    x_3 = arange(center_x, size_x, divisions_size)
    y_3 = arange(center_y, 0, -divisions_size)
    [mesh_x_3, mesh_y_3] = meshgrid(x_3, y_3)

    x_4 = arange(center_x, 0, -divisions_size)
    y_4 = arange(center_y, size_y, divisions_size)
    [mesh_x_4, mesh_y_4] = meshgrid(x_4, y_4)

    [mesh_x, mesh_y] = meshgrid(hstack((x_1, x_2, x_3, x_4)),
                                hstack((y_1, y_2, y_3, y_4)))

    x_horizontal = hstack([mesh_x_1[0, :], mesh_x_2[0, :], mesh_x_3[0, :], mesh_x_4[0, :]])
    y_horizontal = hstack([mesh_y_1[0, :], mesh_y_2[0, :], mesh_y_3[0, :], mesh_y_4[0, :]])

    x_vertical = hstack((mesh_x_1[:, 0], mesh_x_2[:, 0], mesh_x_3[:, 0], mesh_x_4[:, 0]))
    y_vertical = hstack((mesh_y_1[:, 0], mesh_y_2[:, 0], mesh_y_3[:, 0], mesh_y_4[:, 0]))

    return mesh_x, mesh_y, x_horizontal, y_horizontal, x_vertical, y_vertical


# FUNCTIONS FOR THE ASSESSMENT

def anterior_angle_from_horizontal(right_point, left_point, tolerance):
    """
    :param right_point:
    :param left_point:
    :param tolerance:
    :return side_lowered, angle_points:
    """

    distance_points = left_point - right_point
    angle_points = -angle(complex(distance_points[1], distance_points[0]), deg=True)
    angle_points = round(angle_points, 2)

    if angle_points > tolerance:
        side_lowered = 'Der.'
    elif angle_points < -tolerance:
        side_lowered = 'Izq.'
    else:
        side_lowered = 'Alin.'

    return side_lowered, angle_points


def posterior_angle_from_horizontal(left_point, right_point, tolerance):
    """
    :param left_point:
    :param right_point:
    :param tolerance:
    :return side_lowered, angle_points:
    """

    distance_points = right_point - left_point
    angle_points = -angle(complex(distance_points[1], distance_points[0]), deg=True)
    angle_points = round(angle_points, 2)

    if angle_points < tolerance:
        side_lowered = 'Der.'
    elif angle_points > -tolerance:
        side_lowered = 'Izq.'
    else:
        side_lowered = 'Alin.'

    return side_lowered, angle_points


def angle_from_vertical(first_point, second_point, tolerance):
    """
    Anterior view: first point the top, second point the bottom

    Posterior view: first point the bottom, second point the top

    :param first_point:
    :param second_point:
    :param tolerance:
    :return direction, angle_points:
    """

    distance_points = first_point - second_point
    angle_points = 90 - abs(angle(complex(distance_points[1], distance_points[0]), deg=True))
    angle_points = round(angle_points, 2)

    if angle_points < -tolerance:
        direction = 'Izq.'
    elif angle_points > tolerance:
        direction = 'Der.'
    else:
        direction = 'Alin.'

    return direction, angle_points


def lateral_angle_from_vertical(top_point, bottom_point, tolerance):
    """
    :param top_point:
    :param bottom_point:
    :param tolerance:
    :return direction, angle_points:
    """

    distance_points = top_point - bottom_point
    angle_points = 90 - abs(angle(complex(distance_points[1], distance_points[0]), deg=True))
    angle_points = round(angle_points, 2)

    if angle_points < -tolerance:
        direction = 'Pos.'
    elif angle_points > tolerance:
        direction = 'Ant.'
    else:
        direction = 'Alin.'

    return direction, angle_points


def lateral_angle_from_horizontal(posterior_point, anterior_point):
    """
    :param posterior_point:
    :param anterior_point:
    :return direction, angle_points:
    """

    distance_points = posterior_point - anterior_point
    angle_points = 180 - abs(angle(complex(distance_points[1], distance_points[0]), deg=True))
    angle_points = round(angle_points, 2)

    if angle_points > 15:
        direction = 'Ant.'
    elif angle_points < 5:
        direction = 'Pos.'
    else:
        direction = 'Normal'

    return direction, angle_points


def anterior_distance_from_vertical(center_point_x, reference_point, scale, tolerance):
    """
    :param center_point_x:
    :param reference_point:
    :param scale:
    :param tolerance:
    :return direction, distance:
    """

    distance = (center_point_x[1] - reference_point[1]) * scale
    distance = round(distance, 2)

    if distance > tolerance:
        direction = 'Der.'
    elif distance < -tolerance:
        direction = 'Izq.'
    else:
        direction = 'Alin.'

    return direction, distance


def posterior_distance_from_vertical(center_point_x, reference_point, scale, tolerance):
    """
    :param center_point_x:
    :param reference_point:
    :param scale:
    :param tolerance:
    :return direction, distance:
    """

    distance = (reference_point[1] - center_point_x[1]) * scale
    distance = round(distance, 2)

    if distance > tolerance:
        direction = 'Der.'
    elif distance < -tolerance:
        direction = 'Izq.'
    else:
        direction = 'Alin.'

    return direction, distance


def lateral_distance_from_vertical(center_point_x, reference_point, scale, tolerance):
    """
    :param center_point_x:
    :param reference_point:
    :param scale:
    :param tolerance:
    :return direction, distance:
    """

    distance = (reference_point[1] - center_point_x[1]) * scale
    distance = round(distance, 2)

    if distance > tolerance:
        direction = 'Ant.'
    elif distance < -tolerance:
        direction = 'Pos.'
    else:
        direction = 'Alin.'

    return direction, distance


def anterior_rotation(first_point, second_point, tolerance):
    """
    Left side: first point the top, second point the bottom\n
    Right side: first point the bottom, second point the top

    :param first_point:
    :param second_point:
    :param tolerance:
    :return direction, angle_points:
    """

    distance_points = first_point - second_point
    angle_points = abs(angle(complex(distance_points[1], distance_points[0]), deg=True)) - 90
    angle_points = round(angle_points, 2)

    if angle_points > tolerance:
        direction = 'Rot.Ext.'
    elif angle_points < -tolerance:
        direction = 'Rot.Int.'
    else:
        direction = 'Alin.'

    return direction, angle_points


def valgu_or_varus(first_point, second_point, tolerance):
    """
    Anterior:\n
    Left side: first point the bottom, second point the top\n
    Right side: first point the top, second point the bottom

    Posterior:\n
    Left side: first point the top, second point the bottom\n
    Right side: first point the bottom, second point the top

    :param first_point:
    :param second_point:
    :param tolerance:
    :return direction, angle_points:
    """

    distance_points = first_point - second_point
    angle_points = 90 - abs(angle(complex(distance_points[1], distance_points[0]), deg=True))
    angle_points = round(angle_points, 2)

    if angle_points > tolerance:
        direction = 'Valgo'
    elif angle_points < -tolerance:
        direction = 'Varo'
    else:
        direction = 'Alin.'

    return direction, angle_points


# SEGMENTATION

def markers(binary_image):
    """
    :param binary_image:
    :return [reference_1, reference_2, centers_y, centers_x, scale_ratio]:
    """

    binary_image = label(binary_image)
    marker_regions = regionprops(binary_image)
    centers_y = []
    centers_x = []

    for i in range(0, len(marker_regions)):
        y, x = marker_regions[i].centroid
        centers_y.append(y)
        centers_x.append(x)

    marker_centers = [i for i, x in enumerate(centers_x) if x == min(centers_x)]
    reference_1 = [centers_y[marker_centers[0]], centers_x[marker_centers[0]]]
    centers_y.pop(marker_centers[0])
    centers_x.pop(marker_centers[0])

    marker_centers = [i for i, x in enumerate(centers_x) if x == max(centers_x)]
    reference_2 = [centers_y[marker_centers[0]], centers_x[marker_centers[0]]]
    centers_y.pop(marker_centers[0])
    centers_x.pop(marker_centers[0])

    reference_1 = array(reference_1)
    reference_2 = array(reference_2)
    marker_centers_y = array(centers_y)
    marker_centers_x = array(centers_x)
    scale_ratio = 100.0 / sqrt((reference_1[1] ** 2 + reference_2[1] ** 2))

    return reference_1, reference_2, marker_centers_y, marker_centers_x, scale_ratio


# FUNCTION FOR PDF REPORT

class ReportPdf(object):

    def __init__(self, name_pdf):
        super(ReportPdf, self).__init__()

        self.name_pdf = name_pdf
        self.patient_name = ""
        self.styles = getSampleStyleSheet()
        self.width, self.height = A4

    def export_pdf(self, patient, db_dict, img_dict):
        """
        Export the assessment data to PDF.
        
        :param patient: patient's data
        :param db_dict: assessment data dictionary
        :param img_dict: image directory dictionary
        :return:
        """

        global run_anterior_assessment, run_posterior_assessment, run_lateral_r_assessment

        db_anterior = db_dict['Anterior']
        db_posterior = db_dict['Posterior']
        db_lateral_r = db_dict['Lateral_d']

        img_anterior = '.\\final_anterior.jpg'
        img_posterior = '.\\final_posterior.jpg'
        img_lateral = '.\\final_lateral.jpg'

        anterior_img_dir = img_dict['Anterior']
        posterior_img_dir = img_dict['Posterior']
        lateral_r_img_dir = img_dict['Lateral_d']

        copyfile(img_anterior, anterior_img_dir)
        copyfile(img_posterior, posterior_img_dir)
        copyfile(img_lateral, lateral_r_img_dir)

        assessment_date = strftime('%d/%m/%Y')
        self.patient_name = patient.get_name()
        patient_age = patient.get_age()
        patient_gender = patient.get_gender()
        weight_kg = patient.get_weight_kg()
        height_cm = patient.get_height_cm()
        occupation = patient.get_occupation()

        # STYLES FORMAT

        PS = ParagraphStyle

        title_alignment = PS(name='title_alignment', alignment=TA_CENTER, fontSize=14,
                             leading=10, textColor=black,
                             parent=self.styles['Heading1'])

        main_paragraph = PS(name='main_paragraph', alignment=TA_LEFT, fontSize=10,
                            leading=8, textColor=black,
                            parent=self.styles['Heading1'])

        secondary_paragraph = PS(name='secondary_paragraph', alignment=TA_LEFT, fontSize=10,
                                 leading=16, textColor=black)

        data_table_style = (('BACKGROUND', (0, 0), (-1, 0), cornflowerblue),
                            ('TEXTCOLOR', (0, 0), (-1, 0), whitesmoke),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE', black),  # Centered and left-aligned text
                            ('INNERGRID', (0, 0), (-1, -1), 0.50, black),  # Internal lines
                            ('BOX', (0, 0), (-1, -1), 0.25, black),  # External lines
                            )

        results_table_style = (('BACKGROUND', (0, 0), (-1, 0), lightsteelblue),
                               ('TEXTCOLOR', (0, 0), (-1, 0), black),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                               ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE', black),
                               ('INNERGRID', (0, 0), (-1, -1), 0.50, black),
                               ('BOX', (0, 0), (-1, -1), 0.25, black)
                               )

        data_table = Table((('Fecha', 'Nombre', 'Edad', 'Género', 'Talla [cm]', 'Peso [kg]', 'Ocupación'),
                            (assessment_date, self.patient_name, patient_age, patient_gender, height_cm, weight_kg,
                             occupation)), colWidths=(self.width - 100) / 7, hAlign='CENTER', style=data_table_style)

        data_table.argW[1] = 38 * mm  # Modify the cell widt

        lam_img_pdf = None
        try:
            lam_img_pdf = get_image_for_pdf('logo.png', height=100 * mm)
        except OSError:
            print('\nSin archivo "logo.png"')

        record = [Paragraph('Evaluación Postural', title_alignment), Spacer(1, 4 * mm), data_table,
                  lam_img_pdf]  # https://i.imgur.com/KRPTibG.png

        # CONFIGURATION

        conf = read_configuration()
        tolerance_angle = conf['tolerance_angle']
        tolerance_distance = conf['tolerance_distance']

        conf_table = Table((('Ángulo de tolerancia', 'Distancia de toleracia'),
                            (f'{tolerance_angle} °', f'{tolerance_distance} cm')), colWidths=(self.width - 100) / 4,
                           hAlign='CENTER', style=data_table_style)

        record.append(conf_table)

        # DIAGNOSTIC
        if run_anterior_assessment:
            anterior_dg = db_anterior['diagnostic']
            record.append(Spacer(2, 4 * mm))
            record.append(Paragraph('VISTA ANTERIOR:', main_paragraph))
            record.append(Paragraph(f'Línea escapular descendida: {anterior_dg[0]}', secondary_paragraph))
            record.append(Paragraph(f'Línea pélvica descendida: {anterior_dg[1]}', secondary_paragraph))
            record.append(Paragraph(f'Cabeza inclinada: {anterior_dg[2]} con {anterior_dg[3]}°', secondary_paragraph))
            record.append(Paragraph(f'Cabeza rotada: {anterior_dg[4]}', secondary_paragraph))
            record.append(Paragraph(f'Rodillas: Derecha {anterior_dg[5]} {anterior_dg[6]}° ; '
                                    f'Izquierda {anterior_dg[7]} {anterior_dg[8]}°', secondary_paragraph))

        if run_posterior_assessment:
            posterior_dg = db_posterior['diagnostic']
            record.append(Spacer(2, 4 * mm))
            record.append(Paragraph('VISTA POSTERIOR:', main_paragraph))
            record.append(Paragraph(f'Rotación dorsal posterior: {posterior_dg[0]}', secondary_paragraph))
            record.append(Paragraph(f'Rotación pélvica posterior: {posterior_dg[1]}', secondary_paragraph))
            record.append(Paragraph(f'Pie izquierdo: {posterior_dg[2]}', secondary_paragraph))
            record.append(Paragraph(f'Pie derecho: {posterior_dg[3]}', secondary_paragraph))

        if run_lateral_r_assessment:
            lateral_r_dg = db_lateral_r['diagnostic']
            record.append(Spacer(2, 4 * mm))
            record.append(Paragraph('VISTA LATERAL DERECHA:', main_paragraph))
            record.append(Paragraph(f'Postura: {lateral_r_dg[0]}', secondary_paragraph))
            record.append(Paragraph(f'Proyección postural: {lateral_r_dg[1]}', secondary_paragraph))

        record.append(PageBreak())

        # ANTERIOR ASSESSMENT TABLE PAGE

        if run_anterior_assessment:
            df = db_anterior['table_1']

            anterior_table_part_1 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                          colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                          style=results_table_style)

            df = db_anterior['table_2']
            anterior_table_part_2 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                          colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                          style=results_table_style)

            df = db_anterior['table_3']

            anterior_table_part_3 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                          colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                          style=results_table_style)

            record.append(get_image_for_pdf(anterior_img_dir, height=100 * mm))
            record.append(Paragraph('Grados con respecto a la horizontal:', main_paragraph))
            record.append(Paragraph('El ángulo ideal debe ser <strong>0°</strong>.', secondary_paragraph))
            record.append(anterior_table_part_1)
            record.append(Paragraph('Distancia con respecto a la vertical:', main_paragraph))
            record.append(Paragraph('La distancia ideal debe ser <strong>0 cm</strong>.', secondary_paragraph))
            record.append(anterior_table_part_2)
            record.append(Paragraph('Grados de rotación de los pies:', main_paragraph))
            record.append(Paragraph('El ángulo ideal debe ser <strong>0°</strong>.', secondary_paragraph))
            record.append(anterior_table_part_3)
            record.append(PageBreak())

        # POSTERIOR ASSESSMENT TABLE PAGE

        if run_posterior_assessment:
            df = db_posterior['table_1']

            posterior_table_part_1 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            df = db_posterior['table_2']

            posterior_table_part_2 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            df = db_posterior['table_3']

            posterior_table_part_3 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            record.append(get_image_for_pdf(posterior_img_dir, height=100 * mm))
            record.append(Paragraph('Grados con respecto a la horizontal:', main_paragraph))
            record.append(Paragraph('El ángulo ideal debe ser <strong>0°</strong>.', secondary_paragraph))
            record.append(posterior_table_part_1)
            record.append(Paragraph('Distancia con respecto a la vertical:', main_paragraph))
            record.append(Paragraph('La distancia ideal debe ser <strong>0 cm</strong>.', secondary_paragraph))
            record.append(posterior_table_part_2)
            record.append(Paragraph('Grados de rotación de los pies:', main_paragraph))
            record.append(Paragraph('El ángulo ideal debe ser <strong>0°</strong>.', secondary_paragraph))
            record.append(posterior_table_part_3)
            record.append(PageBreak())

        # LATERAL RIGHT ASSESSMENT TABLE PAGE

        if run_lateral_r_assessment:
            df = db_lateral_r['table_1']
            lateral_r_table_part_1 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            df = db_lateral_r['table_2']

            lateral_r_table_part_2 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            df = db_lateral_r['table_3']

            lateral_r_table_part_3 = Table([df.columns.values.tolist()] + df.values.tolist(),
                                           colWidths=(self.width - 100) / 4, hAlign='LEFT',
                                           style=results_table_style)

            record.append(get_image_for_pdf(lateral_r_img_dir, height=100 * mm))
            record.append(Paragraph('Grados con respecto a la vertical:', main_paragraph))
            record.append(Paragraph('El ángulo ideal debe ser <strong>0°</strong>.', secondary_paragraph))
            record.append(lateral_r_table_part_1)
            record.append(Paragraph('Grados con respecto a la horizontal:', main_paragraph))
            record.append(Paragraph('El ángulo normal entre los marcadores pélvicos anterior y posterior <strong>10°'
                                    '</strong>', secondary_paragraph))
            record.append(Paragraph('con una tolerancia de <strong>±5°</strong>.', secondary_paragraph))
            record.append(lateral_r_table_part_2)
            record.append(Paragraph('Distancia con respecto a la vertical:', main_paragraph))
            record.append(Paragraph('La distancia ideal debe ser <strong>0 cm</strong>.', secondary_paragraph))
            record.append(lateral_r_table_part_3)
            record.append(PageBreak())

        file_pdf = SimpleDocTemplate(self.name_pdf, leftMargin=50, rightMargin=50, pagesize=A4, title='Reporte PDF LAM',
                                     author='Youssef Abarca')

        try:
            file_pdf.build(record, onLaterPages=header_and_footer, canvasmaker=PageNumbering)
            record.clear()

            # +------------------------------------+
            return '\nReporte generado con éxito.'
        # +------------------------------------+
        except PermissionError:
            record.clear()
            # +--------------------------------------------+
            return '\nError inesperado: Permiso denegado.'
        # +--------------------------------------------+


def get_image_for_pdf(path_img, height=1 * mm):
    """
    :param path_img:
    :param height:
    :return Image(path, height=height, width=(height * aspect)):
    """

    img = utils.ImageReader(path_img)
    img_width, img_height = img.getSize()
    aspect = img_width / float(img_height)
    return Image(path_img, height=height, width=(height * aspect))


def header_and_footer(canvas_header, file_pdf):
    """
    Save the state of our canvas so that we can take advantage of it

    :param canvas_header:
    :param file_pdf:
    :return:
    """

    canvas_header.saveState()
    style = getSampleStyleSheet()

    alignment = ParagraphStyle(name='alignment', alignment=TA_RIGHT,
                               parent=style['Normal'])

    # HEADER
    header_name = Paragraph(patient_name, style['Normal'])
    header_name.wrap(file_pdf.width, file_pdf.topMargin)
    header_name.drawOn(canvas_header, file_pdf.leftMargin, 270 * mm + (5 * mm))

    date_report = utcnow().to('local').format('dddd, DD - MMMM - YYYY', locale='es')
    date_report = date_report.replace('-', 'de')

    header_date = Paragraph(date_report, alignment)
    header_date.wrap(file_pdf.width, file_pdf.topMargin)
    header_date.drawOn(canvas_header, file_pdf.leftMargin, 270 * mm + (5 * mm))

    # FOOTER
    footer = Paragraph('Lectura Automática de Marcadores', style['Normal'])
    footer.wrap(file_pdf.width, file_pdf.bottomMargin)
    footer.drawOn(canvas_header, file_pdf.leftMargin, 15 * mm + (5 * mm))

    # DROP THE CANVAS
    canvas_header.restoreState()


class PageNumbering(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """
        Add page information to each page (page x of y)
        """

        pages_number = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(pages_number)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.drawRightString(204 * mm, 15 * mm + (5 * mm), f'{self._pageNumber}/{page_count}')


def generate_report(patient, directory, data_dict, img_dict):
    """
    Generate the PDF report with its respective name

    :param patient:
    :param directory:
    :param data_dict:
    :param img_dict:
    :return:
    """

    name_pdf = patient_name + '_' + strftime('%Y%m%d') + '_' + strftime('%H%M%S')
    report = ReportPdf(directory + name_pdf + '.pdf').export_pdf(patient, data_dict, img_dict)
    print(report)
    return name_pdf


# FUNCTIONS FOR POSTURAL ASSESSMENTS

def anterior_assessment(image_dir, tolerance_angle, tolerance_distance, grid_size, my_threshold):
    """
    ANATOMICAL POINTS\n
    f1 = Frown\n
    f2 = Chin\n
    f3 = Right shoulder\n
    f4 = Left shoulder\n
    f5 = Sternum\n
    f6 = Navel\n
    f7 = Right pelvis\n
    f8 = left pelvis\n
    f9 = Right knee\n
    f10 = Left knee\n
    f11 = Right ankle\n
    f12 = Left ankle\n
    f13 = Right big toe\n
    f14 = Left big toe\n
    f_center_shoulders\n
    f_center_y = Center Y (Pelvis)\n
    f_center_knees\n
    f_center_y = Center X (Ankles)\n
    f_center_feet

    :param image_dir:
    :param tolerance_angle:
    :param tolerance_distance:
    :param grid_size:
    :param my_threshold:
    :return:
    """

    markers_num = 14

    img = io.imread(image_dir)
    img = scale_image(img)
    anterior_img = straighten_img(img, my_threshold)

    filtered_anterior_img = green_filter(anterior_img, my_threshold)

    [reference_1, reference_2, marker_centers_y, marker_centers_x, scale_ratio] = markers(filtered_anterior_img)

    if len(marker_centers_x) > markers_num:
        figure(1)
        title(f'Existen demasiados puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(anterior_img)
        show()

    elif len(marker_centers_x) < markers_num:
        figure(1)
        title(f'Existen menos puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(anterior_img)
        show()

    else:

        f1 = (marker_centers_y[0], marker_centers_x[0])
        f2 = (marker_centers_y[1], marker_centers_x[1])

        temp_center_y = [marker_centers_y[2], marker_centers_y[3], marker_centers_y[4]]
        temp_center_x = [marker_centers_x[2], marker_centers_x[3], marker_centers_x[4]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        f3 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        f4 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        f5 = temp_center_y[0], temp_center_x[0]

        temp_center_y = [marker_centers_y[5], marker_centers_y[6], marker_centers_y[7]]
        temp_center_x = [marker_centers_x[5], marker_centers_x[6], marker_centers_x[7]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        f7 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        f8 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        f6 = temp_center_y[0], temp_center_x[0]

        temp_center_y = [marker_centers_y[8], marker_centers_y[9]]
        temp_center_x = [marker_centers_x[8], marker_centers_x[9]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        f9 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        f10 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y = [marker_centers_y[10], marker_centers_y[11]]
        temp_center_x = [marker_centers_x[10], marker_centers_x[11]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        f11 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        f12 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y = [marker_centers_y[12], marker_centers_y[13]]
        temp_center_x = [marker_centers_x[12], marker_centers_x[13]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        f13 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        f14 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        f_center_shoulders = mean([f3, f4], axis=0)
        f_center_y = mean([f7, f8], axis=0)
        f_center_knees = mean([f9, f10], axis=0)
        f_center_x = mean([f11, f12], axis=0)
        f_center_feet = mean([f13, f14], axis=0)

        # F2 y F5 aún sin uso

        f1 = array(f1)
        f2 = array(f2)
        f3 = array(f3)
        f4 = array(f4)
        f5 = array(f5)
        f6 = array(f6)
        f7 = array(f7)
        f8 = array(f8)
        f9 = array(f9)
        f10 = array(f10)
        f11 = array(f11)
        f12 = array(f12)
        f13 = array(f13)
        f14 = array(f14)

        f_center_shoulders = array(f_center_shoulders)
        f_center_y = array(f_center_y)
        f_center_knees = array(f_center_knees)
        f_center_x = array(f_center_x)
        f_center_feet = array(f_center_feet)

        figure(1)
        title('VISTA ANTERIOR')

        plot(marker_centers_x, marker_centers_y, 'b*', markersize="1")

        plot(f_center_shoulders[1], f_center_shoulders[0], 'rx', markersize="3")
        plot(f_center_y[1], f_center_y[0], 'rx', markersize="3")
        plot(f_center_knees[1], f_center_knees[0], 'rx', markersize="3")
        plot(f_center_x[1], f_center_x[0], 'rx', markersize="3")
        plot(f_center_feet[1], f_center_feet[0], 'rx', markersize="3")

        reference_centers = reference_2 - reference_1
        grid_divs = sqrt((reference_centers[0] ** 2 + reference_centers[1] ** 2))
        grid_divs = grid_divs * grid_size / 100

        center_x = f_center_x[1]
        center_y = f_center_y[0]
        size_x = len(anterior_img[1])
        size_y = len(anterior_img)

        (mesh_x, mesh_y, x_horizontal, y_horizontal, x_vertical, y_vertical) = \
            img_grid(center_x, center_y, size_x, size_y, grid_divs)

        plot(mesh_x, mesh_y, 'k', mesh_x.T, mesh_y.T, 'k', linewidth=0.1)
        plot(x_horizontal, y_horizontal, 'r', linewidth=0.3)
        plot(x_vertical, y_vertical, 'r', linewidth=0.3)

        io.imshow(anterior_img)

        anterior_img_dir = '.\\final_anterior.jpg'

        savefig(anterior_img_dir, dpi=500)
        # time.sleep(1)
        close()  # To show the image change "close" to "show"

        # TA1
        shoulder_low, shoulder_angle = anterior_angle_from_horizontal(f3, f4, tolerance_angle)
        pelvis_low, pelvis_angle = anterior_angle_from_horizontal(f7, f8, tolerance_angle)
        knee_low, knee_angle = anterior_angle_from_horizontal(f9, f10, tolerance_angle)

        # TA2
        forehead_dir, forehead_dis = \
            anterior_distance_from_vertical(f_center_x, f1, scale_ratio, tolerance_distance)
        shoulders_dir, shoulders_dis = \
            anterior_distance_from_vertical(f_center_x, f_center_shoulders, scale_ratio, tolerance_distance)
        navel_dir, navel_dis = \
            anterior_distance_from_vertical(f_center_x, f6, scale_ratio, tolerance_distance)
        pelvis_dir, pelvis_dis = \
            anterior_distance_from_vertical(f_center_x, f_center_y, scale_ratio, tolerance_distance)
        knees_dir, knees_dis = \
            anterior_distance_from_vertical(f_center_x, f_center_knees, scale_ratio, tolerance_distance)
        feet_dir, feet_dis = \
            anterior_distance_from_vertical(f_center_x, f_center_feet, scale_ratio, tolerance_distance)

        # TA3
        left_foot_rot, left_foot_angle = anterior_rotation(f12, f14, tolerance_angle)
        right_foot_rot, right_foot_angle = anterior_rotation(f13, f11, tolerance_angle)

        # DIAGNOSTIC

        head_tilt, head_angle = angle_from_vertical(f1, f2, tolerance_angle)

        if head_tilt == 'Alin.':
            head_tilt = 'N/A'

        head_rot, _ = angle_from_vertical(f2, f5, tolerance_angle)

        if head_rot == 'Alin.':
            head_rot = 'N/A'

        left_knee_dir, left_knee_angle = valgu_or_varus(f12, f10, tolerance_angle)
        right_knee_dir, right_knee_angle = valgu_or_varus(f9, f11, tolerance_angle)

        data_diagnostic = (shoulder_low, pelvis_low, head_tilt, head_angle, head_rot, right_knee_dir, right_knee_angle,
                           left_knee_dir, left_knee_angle)

        data_table_1 = {'Segmento Corporal': ('Hombros', 'Pelvis', 'Rodillas'),
                        'Descendido': (shoulder_low, pelvis_low, knee_low),
                        'Ángulo': (shoulder_angle, pelvis_angle, knee_angle)}

        df_table_1 = DataFrame(data_table_1)

        data_table_2 = {'Referencia': ('Frente', 'Hombros', 'Ombligo', 'Pelvis', 'Rodillas', 'Pies'),
                        'Dirección': (forehead_dir, shoulders_dir, navel_dir, pelvis_dir,
                                      knees_dir, feet_dir),
                        'Distancia': (forehead_dis, shoulders_dis, navel_dis, pelvis_dis,
                                      knees_dis, feet_dis)}

        df_table_2 = DataFrame(data_table_2)

        data_table_3 = {'Segmento Corporal': ('Pie Izquierdo', 'Pie Derecho'),
                        'Dirección': (left_foot_rot, right_foot_rot),
                        'Ángulo': (left_foot_angle, right_foot_angle)}

        df_table_3 = DataFrame(data_table_3)

        # PRINT RESULTS

        print(f'\n\n{"*" * 10} VISTA ANTERIOR {"*" * 10}\n')

        print(f'Línea escapular descendida: {shoulder_low}')
        print(f'Línea pélvica descendida: {pelvis_low}')
        print(f'Cabeza inclinada: {head_tilt} con {head_angle}°')
        print(f'Cabeza rotada: {head_rot}')
        print(f'Rodillas: Derecha {right_knee_dir} {right_knee_angle}° ; '
              f'Izquierda {left_knee_dir} {left_knee_angle}°')

        print('\nGRADOS CON RESPECTO A LA HORIZONTAL:')
        print('El ángulo ideal debe ser 0°.\n')
        print(df_table_1)
        print('\n\nDISTANCIA CON RESPECTO A LA VERTICAL:')
        print('La distancia ideal debe ser 0 cm.\n')
        print(df_table_2)
        print('\n\nGRADOS DE ROTACIÓN DE LOS PIES:\n')
        print(df_table_3)

        datos = {'diagnostic': data_diagnostic,
                 'table_1': df_table_1,
                 'table_2': df_table_2,
                 'table_3': df_table_3}

        show_image = input('\nMostrar image 1/0: ').strip() == '1'
        if show_image:
            open_new(anterior_img_dir)
            input('Presione una tecla para continuar... ')

        return datos
    return 0


def posterior_assessment(image_dir, tolerance_angle, tolerance_distance, grid_size, my_threshold):
    """
    ANATOMICAL POINTS\n
    p1 = 7th Cervical\n
    p2 = 5th Thoracic\n
    p3 = Right shoulder\n
    p4 = Left shoulder\n
    p5 = Right pelvis\n
    p6 = Left pelvis\n
    p7 = Right knee\n
    p8 = Left knee\n
    p9 = Right ankle\n
    p10 = Left ankle\n
    p11 = Right sole\n
    p12 = Left sole\n
    p_center_shoulders\n
    p_center_y = Center Y (Pelvis)\n
    p_center_knees\n
    p_center_ankles\n
    p_center_x = Center X (Soles)

    :param image_dir:
    :param tolerance_angle:
    :param tolerance_distance:
    :param grid_size:
    :param my_threshold:
    :return:
    """

    markers_num = 12

    img = io.imread(image_dir)
    img = scale_image(img)
    posterior_img = straighten_img(img, my_threshold)

    filtered_posterior_img = green_filter(posterior_img, my_threshold)

    [reference_1, reference_2, marker_centers_y, marker_centers_x, scale_ratio] = markers(filtered_posterior_img)

    if len(marker_centers_x) > markers_num:
        figure(1)
        title(f'Existen demasiados puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(posterior_img)
        show()

    elif len(marker_centers_x) < markers_num:
        figure(1)
        title(f'Existen menos puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(posterior_img)
        show()

    else:

        temp_center_y = [marker_centers_y[0], marker_centers_y[1], marker_centers_y[2]]
        temp_center_x = [marker_centers_x[0], marker_centers_x[1], marker_centers_x[2]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        p4 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        p3 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y.pop(pos[0])
        temp_center_x.pop(pos[0])
        p1 = temp_center_y[0], temp_center_x[0]

        p2 = (marker_centers_y[3], marker_centers_x[3])

        temp_center_y = [marker_centers_y[4], marker_centers_y[5]]
        temp_center_x = [marker_centers_x[4], marker_centers_x[5]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        p6 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        p5 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y = [marker_centers_y[6], marker_centers_y[7]]
        temp_center_x = [marker_centers_x[6], marker_centers_x[7]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        p8 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        p7 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y = [marker_centers_y[8], marker_centers_y[9]]
        temp_center_x = [marker_centers_x[8], marker_centers_x[9]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        p10 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        p9 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        temp_center_y = [marker_centers_y[10], marker_centers_y[11]]
        temp_center_x = [marker_centers_x[10], marker_centers_x[11]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        p12 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        p11 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        p_center_shoulders = mean([p4, p3], axis=0)
        p_center_y = mean([p6, p5], axis=0)
        p_center_knees = mean([p8, p7], axis=0)
        p_center_ankles = mean([p10, p9], axis=0)
        p_center_x = mean([p12, p11], axis=0)

        p1 = array(p1)
        p2 = array(p2)
        p3 = array(p3)
        p4 = array(p4)
        p5 = array(p5)
        p6 = array(p6)
        p7 = array(p7)
        p8 = array(p8)
        p9 = array(p9)
        p10 = array(p10)
        p11 = array(p11)
        p12 = array(p12)

        p_center_shoulders = array(p_center_shoulders)
        p_center_y = array(p_center_y)
        p_center_knees = array(p_center_knees)
        p_center_ankles = array(p_center_ankles)
        p_center_x = array(p_center_x)

        figure(1)
        title('VISTA POSTERIOR')

        plot(marker_centers_x, marker_centers_y, 'b*', markersize="1")

        plot(p_center_shoulders[1], p_center_shoulders[0], 'rx', markersize="3")
        plot(p_center_y[1], p_center_y[0], 'rx', markersize="3")
        plot(p_center_knees[1], p_center_knees[0], 'rx', markersize="3")
        plot(p_center_ankles[1], p_center_ankles[0], 'rx', markersize="3")
        plot(p_center_x[1], p_center_x[0], 'rx', markersize="3")

        reference_centers = reference_2 - reference_1
        grid_divs = sqrt((reference_centers[0] ** 2 + reference_centers[1] ** 2))
        grid_divs = grid_divs * grid_size / 100

        center_x = p_center_x[1]
        center_y = p_center_y[0]
        size_x = len(posterior_img[1])
        size_y = len(posterior_img)

        (mesh_x, mesh_y, x_horizontal, y_horizontal, x_vertical, y_vertical) = \
            img_grid(center_x, center_y, size_x, size_y, grid_divs)

        plot(mesh_x, mesh_y, 'k', mesh_x.T, mesh_y.T, 'k', linewidth=0.1)
        plot(x_horizontal, y_horizontal, 'r', linewidth=0.3)
        plot(x_vertical, y_vertical, 'r', linewidth=0.3)

        io.imshow(posterior_img)

        posterior_img_dir = '.\\final_posterior.jpg'

        savefig(posterior_img_dir, dpi=500)
        # time.sleep(1)
        close()  # To show the image change "close" to "show"

        # TP1

        shoulder_low, shoulder_angle = posterior_angle_from_horizontal(p4, p3, tolerance_angle)
        pelvis_low, pelvis_angle = posterior_angle_from_horizontal(p6, p5, tolerance_angle)
        knee_low, knee_angle = posterior_angle_from_horizontal(p8, p7, tolerance_angle)

        # TP2
        shoulders_dir, shoulders_dis = \
            posterior_distance_from_vertical(p_center_x, p_center_shoulders, scale_ratio, tolerance_distance)
        seventh_cervical_dir, seventh_cervical_dis = \
            posterior_distance_from_vertical(p_center_x, p1, scale_ratio, tolerance_distance)
        fifth_thoracic_dir, fifth_thoracic_dis = \
            posterior_distance_from_vertical(p_center_x, p2, scale_ratio, tolerance_distance)
        pelvis_dir, pelvis_dis = \
            posterior_distance_from_vertical(p_center_x, p_center_y, scale_ratio, tolerance_distance)
        knees_dir, knees_dis = \
            posterior_distance_from_vertical(p_center_x, p_center_knees, scale_ratio, tolerance_distance)
        ankles_dir, ankles_dis = \
            posterior_distance_from_vertical(p_center_x, p_center_ankles, scale_ratio, tolerance_distance)

        # TP3
        left_foot_dir, left_foot_angle = valgu_or_varus(p10, p12, tolerance_angle)
        right_foot_dir, right_foot_angle = valgu_or_varus(p11, p9, tolerance_angle)

        # DIAGNOSTIC

        dorsal_rot, _ = angle_from_vertical(p_center_y, p_center_shoulders, tolerance_angle)

        pelvis_rot, _ = angle_from_vertical(p_center_knees, p_center_y, tolerance_angle)

        data_diagnostic = (dorsal_rot, pelvis_rot, left_foot_dir, right_foot_dir)

        data_table_1 = {'Segmento Corporal': ('Hombros', 'Pelvis', 'Rodillas'),
                        'Descendido': (shoulder_low, pelvis_low, knee_low),
                        'Ángulo': (shoulder_angle, pelvis_angle, knee_angle)}

        df_table_1 = DataFrame(data_table_1)

        data_table_2 = {'Referencia': ('Frente', 'Hombros', 'Ombligo', 'Pelvis', 'Rodillas', 'Pies'),
                        'Dirección': (shoulders_dir, seventh_cervical_dir, fifth_thoracic_dir,
                                      pelvis_dir, knees_dir, ankles_dir),
                        'Distancia': (shoulders_dis, seventh_cervical_dis, fifth_thoracic_dis,
                                      pelvis_dis, knees_dis, ankles_dis)}

        df_table_2 = DataFrame(data_table_2)

        data_table_3 = {'Segmento Corporal': ('Pie Izquierdo', 'Pie Derecho'),
                        'Dirección': (left_foot_dir, right_foot_dir),
                        'Ángulo': (left_foot_angle, right_foot_angle)}

        df_table_3 = DataFrame(data_table_3)

        print(f'\n\n{"*" * 10} VISTA POSTERIOR {"*" * 10}\n')

        print(f'Rotación dorsal posterior: {dorsal_rot}')
        print(f'Rotación pélvica posterior: {pelvis_rot}')
        print(f'Pie izquierdo: {left_foot_dir}')
        print(f'Pie derecho: {right_foot_dir}')

        print('\nGRADOS CON RESPECTO A LA HORIZONTAL:')
        print('El ángulo ideal debe ser 0°.\n')
        print(df_table_1)
        print('\n\nDISTANCIA CON RESPECTO A LA VERTICAL:')
        print('La distancia ideal debe ser 0 cm.\n')
        print(df_table_2)
        print('\n\nGRADOS DE ROTACIÓN DE LOS PIES:')
        print('El ángulo ideal debe ser 0°.\n')
        print(df_table_3)

        datos = {'diagnostic': data_diagnostic,
                 'table_1': df_table_1,
                 'table_2': df_table_2,
                 'table_3': df_table_3}

        show_image = input('\nMostrar image 1/0: ').strip() == '1'
        if show_image:
            open_new(posterior_img_dir)
            input('Presione una tecla para continuar... ')

        return datos
    return 0


def lateral_r_assessment(image_directory, tolerance_angle, tolerance_distance, grid_size, my_threshold):
    """
    ANATOMICAL POINTS\n
    l1 = Ear\n
    l2 = Shoulder\n
    l3 = Posterior pelvis\n
    l4 = Anterior pelvis\n
    l5 = Hip\n
    l6 = Knee\n
    l7 = Sole\n
    l_center_y = Center Y (Pelvis)

    :param image_directory:
    :param tolerance_angle:
    :param tolerance_distance:
    :param grid_size:
    :param my_threshold:
    :return:
    """
    markers_num = 7

    img = io.imread(image_directory)
    img = scale_image(img)
    lateral_r_img = straighten_img(img, my_threshold)

    filtered_lateral_r_img = green_filter(lateral_r_img, my_threshold)

    [reference_1, reference_2, marker_centers_y, marker_centers_x, scale_ratio] = markers(filtered_lateral_r_img)

    if len(marker_centers_x) > markers_num:
        figure(1)
        title(f'Existen demasiados puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(lateral_r_img)
        show()

    elif len(marker_centers_x) < markers_num:
        figure(1)
        title(f'Existen menos puntos en la imagen:\n{len(marker_centers_x)} puntos de {markers_num}')
        plot(marker_centers_x, marker_centers_y, 'b*', markersize="5")
        io.imshow(lateral_r_img)
        show()

    else:

        l1 = (marker_centers_y[0], marker_centers_x[0])
        l2 = (marker_centers_y[1], marker_centers_x[1])

        temp_center_y = [marker_centers_y[2], marker_centers_y[3]]
        temp_center_x = [marker_centers_x[2], marker_centers_x[3]]
        pos = [i for i, x in enumerate(temp_center_x) if x == min(temp_center_x)]
        l3 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        pos = [i for i, x in enumerate(temp_center_x) if x == max(temp_center_x)]
        l4 = temp_center_y[pos[0]], temp_center_x[pos[0]]

        l5 = (marker_centers_y[4], marker_centers_x[4])
        l6 = (marker_centers_y[5], marker_centers_x[5])
        l7 = (marker_centers_y[6], marker_centers_x[6])

        l_center_y = mean([l3, l4], axis=0)

        l1 = array(l1)
        l2 = array(l2)
        l3 = array(l3)
        l4 = array(l4)
        l5 = array(l5)
        l6 = array(l6)
        l7 = array(l7)

        l_center_y = array(l_center_y)

        figure(1)
        title('VISTA LATERAL DERECHA')

        plot(marker_centers_x, marker_centers_y, 'b*', markersize="1")

        plot(l_center_y[1], l_center_y[0], 'rx', markersize="3")

        reference_centers = reference_2 - reference_1
        grid_divs = sqrt((reference_centers[0] ** 2 + reference_centers[1] ** 2))
        grid_divs = grid_divs * grid_size / 100

        center_x = l7[1]
        center_y = l_center_y[0]
        size_x = len(lateral_r_img[1])
        size_y = len(lateral_r_img)

        (mesh_x, mesh_y, x_horizontal, y_horizontal, x_vertical, y_vertical) = \
            img_grid(center_x, center_y, size_x, size_y, grid_divs)

        plot(mesh_x, mesh_y, 'k', mesh_x.T, mesh_y.T, 'k', linewidth=0.1)
        plot(x_horizontal, y_horizontal, 'r', linewidth=0.3)
        plot(x_vertical, y_vertical, 'r', linewidth=0.3)

        io.imshow(lateral_r_img)

        lateral_r_img_dir = '.\\final_lateral.jpg'

        savefig(lateral_r_img_dir, dpi=500)
        # time.sleep(1)
        close()  # To show the image change "close" to "show"

        # TL1
        head_shoulder_dir, head_shoulder_angle = lateral_angle_from_vertical(l1, l2, tolerance_angle)
        shoulder_pelvis_dir, shoulder_pelvis_angle = lateral_angle_from_vertical(l2, l_center_y, tolerance_angle)
        hip_knee_dir, hip_knee_angle = lateral_angle_from_vertical(l5, l6, tolerance_angle)
        knee_foot_dir, knee_foot_angle = lateral_angle_from_vertical(l6, l7, tolerance_angle)

        # TL2
        pelvis_dir, pelvis_angle = lateral_angle_from_horizontal(l3, l4)

        # TL3
        head_dir, head_dis = lateral_distance_from_vertical(l7, l1, scale_ratio, tolerance_distance)
        shoulder_dir, shoulder_dis = lateral_distance_from_vertical(l7, l2, scale_ratio, tolerance_distance)
        pelvis_dir_table_3, pelvis_dis = lateral_distance_from_vertical(l7, l_center_y, scale_ratio, tolerance_distance)
        hip_dir, hip_dis = lateral_distance_from_vertical(l7, l5, scale_ratio, tolerance_distance)
        knee_dir, knee_dis = lateral_distance_from_vertical(l7, l6, scale_ratio, tolerance_distance)

        # DIAGNOSTIC

        if pelvis_dir == 'Normal' and head_dis <= 0.5 and shoulder_dis <= 0.5 \
                and hip_dis <= 0.5 and knee_dis <= 0.5:
            posture = 'Ideal'

        elif head_shoulder_angle > 13 and head_dir == 'Ant.' and shoulder_pelvis_dir != 'Ant.' \
                and pelvis_dir == 'Ant.':
            posture = 'Cifolordótica'

        elif head_dir == 'Ant.' and shoulder_pelvis_dir != 'Pos.' and pelvis_dir_table_3 != 'Pos.' \
                and pelvis_dir == 'Pos.':
            posture = 'Espalda plana'

        elif head_dir == 'Ant.' and shoulder_pelvis_dir == 'Pos.' and pelvis_dir_table_3 != 'Pos.' \
                and pelvis_dir == 'Pos.':
            posture = 'Espalda arqueada'

        elif pelvis_dir == 'Normal' and head_dir == 'Ant.' and shoulder_pelvis_dir != 'Ant.':
            posture = 'Cabeza hacia delante'

        elif pelvis_dir == 'Ant.' and shoulder_pelvis_dir != 'Ant.':
            posture = 'Lordótica o tipo militar'

        else:
            if pelvis_dir == 'Ant.':
                posture = 'Pelvis en Anteversión'
            elif pelvis_dir == 'Pos.':
                posture = 'Pelvis en Retroversión'
            else:
                posture = 'Pelvis en posición normal'

        if head_dir == 'Alin.':
            postural_projection = 'Centrada'
        elif head_dir == 'Ant.':
            postural_projection = 'Anterior'
        else:
            postural_projection = 'Posterior'

        data_diagnostic = (posture, postural_projection)

        data_table_1 = {'Segmento Corporal': ('Cabeza-Hombro', 'Hombro-Pelvis', 'Cadera-Rodilla', 'Rodilla-Pies'),
                        'Dirección': (head_shoulder_dir, shoulder_pelvis_dir, hip_knee_dir,
                                      knee_foot_dir),
                        'Ángulo': (head_shoulder_angle, shoulder_pelvis_angle, hip_knee_angle,
                                   knee_foot_angle)}

        df_table_1 = DataFrame(data_table_1)

        data_table_2 = {'Segmento Corporal': ['Pelvis'],
                        'Dirección': [pelvis_dir],
                        'Ángulo': [pelvis_angle]}

        df_table_2 = DataFrame(data_table_2)

        data_table_3 = {'Referencia': ('Cabeza', 'Hombros', 'Pelvis', 'Caderas', 'Rodillas'),
                        'Dirección': (head_dir, shoulder_dir, pelvis_dir_table_3, hip_dir,
                                      knee_dir),
                        'Distancia': (head_dis, shoulder_dis, pelvis_dis, hip_dis,
                                      knee_dis)}

        df_table_3 = DataFrame(data_table_3)

        print(f'\n\n{"*" * 10} VISTA LATERAL DERECHA {"*" * 10}\n')
        print(f'Postura: {posture}')
        print(f'Proyección postural: {postural_projection}')

        print('\nGRADOS CON RESPECTO A LA VERTICAL:')
        print('El ángulo ideal debe ser 0°.\n')
        print(df_table_1)
        print('\n\nGRADOS CON RESPECTO A LA HORIZONTAL:')
        print('El ángulo normal entre los marcadores pélvicos anterior y posterior')
        print('de 10° con una tolerancia de ±5°.\n')
        print(df_table_2)
        print('\n\nDISTANCIA CON RESPECTO A LA VERTICAL:')
        print('La distancia ideal debe ser 0 cm.\n')
        print(df_table_3)

        datos = {'diagnostic': data_diagnostic,
                 'table_1': df_table_1,
                 'table_2': df_table_2,
                 'table_3': df_table_3}

        show_image = input('\nMostrar image 1/0: ').strip() == '1'
        if show_image:
            open_new(lateral_r_img_dir)
            input('Presione una tecla para continuar... ')

        return datos
    return 0


def header_table_format(sheet_name, text, start_cell, end_cell):
    """
    :param sheet_name:
    :param text: Cell Value
    :param start_cell:
    :param end_cell:
    :return:
    """

    thick = Side(border_style='thick', color='000000')

    sheet_name.merge_cells(f'{start_cell}:{end_cell}')
    cell = sheet_name[start_cell]
    cell.value = text

    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True, italic=True)
    cell.fill = PatternFill('solid', fgColor='00C0C0C0')

    sheet_name.merge_cells(f'{start_cell}:{end_cell}')
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)


def create_xlsx(db_xlsx_dir):
    writer = ExcelWriter(db_xlsx_dir, engine='openpyxl')
    DataFrame().to_excel(writer, 'Anterior', index=False)
    DataFrame().T.to_excel(writer, 'Posterior', index=False)
    DataFrame().T.to_excel(writer, 'LateralD', index=False)
    writer.save()
    writer.close()

    wb = load_workbook(db_xlsx_dir)
    wb.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    name_sheets = list(writer.sheets.keys())

    # ANTERIOR

    sheet = wb[name_sheets[0]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'O1')
    header_table_format(sheet, 'Grados con la horizontal', 'P1', 'R1')
    header_table_format(sheet, 'Distancia con la vertical', 'S1', 'X1')
    header_table_format(sheet, 'Rotacion de los pies', 'Y1', 'Z1')

    # POSTERIOR

    sheet = wb[name_sheets[1]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'M1')
    header_table_format(sheet, 'Grados con la horizontal', 'N1', 'P1')
    header_table_format(sheet, 'Distancia con la vertical', 'Q1', 'V1')
    header_table_format(sheet, 'Rotacion de los pies', 'W1', 'X1')

    # LATERAL R

    sheet = wb[name_sheets[2]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'K1')
    header_table_format(sheet, 'Grados con la vertical', 'L1', 'O1')
    header_table_format(sheet, 'Grados con la horizontal', 'P1', 'P1')
    header_table_format(sheet, 'Distancia con la vertical', 'Q1', 'U1')

    wb.save(db_xlsx_dir)
    wb.close()

    with ExcelWriter(db_xlsx_dir, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        anterior_header = DataFrame((), ('Imagen', 'Reporte',

                                         'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                         'Linea escapular descendida', 'Linea pélvica descendida',
                                         'Cabeza inclinada', 'Cabeza rotada', 'Rodilla derecha', 'Rodilla izquierda',

                                         'Hombro descendido', 'Pelvis descendida', 'Rodilla descendida',

                                         'Frente', 'Hombros', 'Ombligo', 'Pelvis', 'Rodillas', 'Pies',

                                         'Pie izquierdo', 'Pie derecho'))

        posterior_header = DataFrame((), ('Imagen', 'Reporte',

                                          'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                          'Rotación dorsal posterior', 'Rotación pélvica posterior', 'Pie izquierdo',
                                          'Pie derecho',

                                          'Hombro descendido', 'Pelvis descendida', 'Rodilla descendida',

                                          'Hombros', '7ma cervical', '5ta torácica', 'Pelvis', 'Rodillas',
                                          'Tobillos',

                                          'Pie izquierdo', 'Pie derecho',))

        lateral_r_header = DataFrame((), ('Imagen', 'Rreporte',

                                          'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                          'Postura', 'Proyección postural',

                                          'Cabeza-hombro', 'Hombro-pelvis', 'Cadera-rodilla', 'Rodilla-pie',

                                          'Pelvis',

                                          'Cabeza', 'Hombro', 'Pelvis', 'Cadera', 'Rodilla'))

        anterior_header.T.to_excel(writer, name_sheets[0], header=True, index=False, startrow=1, startcol=0)
        posterior_header.T.to_excel(writer, name_sheets[1], header=True, index=False, startrow=1, startcol=0)
        lateral_r_header.T.to_excel(writer, name_sheets[2], header=True, index=False, startrow=1, startcol=0)


def transform_information_xlsx(img_dir, pdf_dir, patient_data, diagnostic, db_assessment):
    df1 = db_assessment['table_1']
    df2 = db_assessment['table_2']
    df3 = db_assessment['table_3']
    values_tables = df1.values.tolist() + df2.values.tolist() + df3.values.tolist()
    values_tables = [i[:][2] for i in values_tables]

    img_dir = [f'=HYPERLINK(\"{img_dir}\",\"{patient_data[1]}  {patient_data[0]}\")']

    data = img_dir + pdf_dir + patient_data + diagnostic + values_tables
    data = DataFrame(data)

    return data.T


def add_data_db_xlsx(db_xlsx_dir, patient, data_dict, img_dict, report_pdf_dir):
    if not path.isfile(db_xlsx_dir):
        create_xlsx(db_xlsx_dir)

    assessment_date = strftime('%d/%m/%Y')
    db_anterior = data_dict['Anterior']
    db_posterior = data_dict['Posterior']
    db_lateral_r = data_dict['Lateral_d']

    anterior_img_dir = img_dict['Anterior']
    posterior_img_dir = img_dict['Posterior']
    lateral_r_img_dir = img_dict['Lateral_d']

    patient_age = patient.get_age()
    patient_gender = patient.get_gender()
    weight_kg = patient.get_weight_kg()
    height_cm = patient.get_height_cm()
    occupation = patient.get_occupation()

    patient_data = [assessment_date, patient_name, patient_age, patient_gender, height_cm, weight_kg, occupation]
    report_pdf_dir = [f'=HYPERLINK(\"{report_pdf_dir}\",\"{patient_name}  {assessment_date}\")']
    wb = load_workbook(db_xlsx_dir)

    try:

        with ExcelWriter(db_xlsx_dir, engine='openpyxl') as writer:

            writer.book = wb
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            name_sheets = list(writer.sheets.keys())

            if run_anterior_assessment:
                sheet = wb[name_sheets[0]]
                max_row = sheet.max_row

                dg = db_anterior['diagnostic']
                dg = [dg[0], dg[1], dg[3], dg[4], dg[6], dg[8]]

                data = transform_information_xlsx(anterior_img_dir, report_pdf_dir, patient_data, dg, db_anterior)
                data.to_excel(writer, name_sheets[0], header=False, index=False, startrow=max_row, startcol=0)

            if run_posterior_assessment:
                sheet = wb[name_sheets[1]]
                max_row = sheet.max_row

                dg = db_posterior['diagnostic']
                dg = [dg[0], dg[1], dg[2], dg[3]]

                data = transform_information_xlsx(posterior_img_dir, report_pdf_dir, patient_data, dg, db_posterior)
                data.to_excel(writer, name_sheets[1], header=False, index=False, startrow=max_row, startcol=0)

            if run_lateral_r_assessment:
                sheet = wb[name_sheets[2]]
                max_row = sheet.max_row

                dg = db_posterior['diagnostic']
                dg = [dg[0], dg[1]]

                data = transform_information_xlsx(lateral_r_img_dir, report_pdf_dir, patient_data, dg, db_lateral_r)
                data.to_excel(writer, name_sheets[2], header=False, index=False, startrow=max_row, startcol=0)

        wb.close()
        print('\nDatos agregados con éxito')

    except PermissionError:
        wb.close()
        print('\nError de escritura')
        print('Revise que el archivo no se encuentre abierto o los permisos del antivirus\n')
        reload_data = bool(int(input("¿Desea cargar nuevamente? 1/0: ")))

        if reload_data:
            add_data_db_xlsx(db_xlsx_dir, patient, data_dict, img_dict, report_pdf_dir)


def postural_assessment(tolerance_angle, tolerance_distance, grid_size, my_threshold):
    """
    :param tolerance_angle:
    :param tolerance_distance:
    :param grid_size:
    :param my_threshold:
    :return:
    """

    global photo
    global run_anterior_assessment, run_posterior_assessment, run_lateral_r_assessment

    global patient_name

    print('Parametros de configuración:\n')
    print(f'El ángulo de toleracia es: {tolerance_angle} °')
    print(f'La distancia de toleracia es: {tolerance_distance} cm')
    print(f'Tamaño de la cuadrícula: {grid_size} cm')
    print(f'Valor umbral del filtro: {my_threshold}')

    # What assessment will be performed

    print("\nEvaluaciones a realizar => 1: Si | 0: NO\n")

    run_anterior_assessment = input('Vista Anterior 1/0: ').strip() == '1'
    run_posterior_assessment = input('Vista Posterior 1/0: ').strip() == '1'
    run_lateral_r_assessment = input('Vista Lateral Derecha 1/0: ').strip() == '1'

    anterior_results = 0
    posterior_results = 0
    lateral_r_results = 0

    # Image upload test in case of wrong images

    if run_anterior_assessment:
        while anterior_results == 0:
            photo = ''
            try:
                load_image("\"Anterior\"")  # https://i.imgur.com/qRb4dv6.jpg
                anterior_results = anterior_assessment(photo, tolerance_angle, tolerance_distance, grid_size,
                                                       my_threshold)
            except ValueError:
                reload_img = input('Imagen Anterior con Error. ¿Desea cargar otra image? 1/0: ').strip() == '1'
                if not reload_img:
                    run_anterior_assessment = False
                    break

    if run_posterior_assessment:
        while posterior_results == 0:
            photo = ''
            try:
                load_image("\"Posterior\"")  # https://i.imgur.com/xIYYjkc.jpg
                posterior_results = posterior_assessment(photo, tolerance_angle, tolerance_distance, grid_size,
                                                         my_threshold)
            except ValueError:
                reload_img = input('Imagen Posterior con Error. ¿Desea cargar otra image? 1/0: ').strip() == '1'
                if not reload_img:
                    run_posterior_assessment = False
                    break

    if run_lateral_r_assessment:
        while lateral_r_results == 0:
            photo = ''
            try:
                load_image("\"Lareral derecha\"")  # https://i.imgur.com/2fvjwk1.jpg
                lateral_r_results = lateral_r_assessment(photo, tolerance_angle, tolerance_distance, grid_size,
                                                         my_threshold)
            except ValueError:
                reload_img = input("Imagen Lateral Derecha con Error. ¿Desea cargar otra image? 1/0: ").strip() == '1'
                if not reload_img:
                    run_lateral_r_assessment = False
                    break

    if run_anterior_assessment or run_posterior_assessment or run_lateral_r_assessment:
        generate_pdf = input("Desea generar reporte 1/0: ").strip() == '1'
    else:
        generate_pdf = False

    if generate_pdf:

        # Personal information

        print("\nDATOS PERSONALES")
        patient = Patient()

        patient_name = patient.get_name()

        patient_dir, images_dir, db_xlsx_dir = create_directories()

        anterior_name_img = 'Anterior' + '_' + strftime("%Y%m%d") + '_' + strftime("%H%M%S") + '.jpg'
        posterior_name_img = 'Posterior' + '_' + strftime("%Y%m%d") + '_' + strftime("%H%M%S") + '.jpg'
        lateral_r_name_img = 'LateralD' + '_' + strftime("%Y%m%d") + '_' + strftime("%H%M%S") + '.jpg'

        data_dict = {'Anterior': anterior_results,
                     'Posterior': posterior_results,
                     'Lateral_d': lateral_r_results}

        img_dict = {'Anterior': images_dir + anterior_name_img,
                    'Posterior': images_dir + posterior_name_img,
                    'Lateral_d': images_dir + lateral_r_name_img}

        name_pdf = generate_report(patient, patient_dir, data_dict, img_dict)

        report_pdf_dir = patient_dir + name_pdf + '.pdf'

        input('Presione una tecla para continuar... ')
        open_new(report_pdf_dir)

        ###############################

        add_data_db = input("Desea agregar datos a la base 1/0: ").strip() == '1'

        if add_data_db:
            add_data_db_xlsx(db_xlsx_dir, patient, data_dict, img_dict, report_pdf_dir)

    input('Presione una tecla para continuar... ')
    system("cls")


# OTHER FUNCTIONS

class Patient:

    def __init__(self):

        self.name = input("Nombre: ")
        self.name = self.name.strip().title()
        if self.name == '':
            self.name = 'Anónimo'

        try:
            self.age = int(input("Edad: "))
            while self.age <= 0:
                print('Ingrese un valor válido > 0')
                self.age = int(input("Edad: "))
        except ValueError:
            print('Error en ingreso de edad')
            self.age = None

        self.gender = input("Género (M/F): ")
        self.gender = self.gender.strip().upper()
        if not (self.gender == 'M' or self.gender == 'F'):
            print('Error al ingresar género')
            self.gender = None

        try:
            self.height_cm = float(input("Altura en cm: "))
            while self.height_cm <= 0:
                print('Ingrese un valor válido > 0')
                self.height_cm = float(input("Altura en cm: "))
        except ValueError:
            print('Error en ingreso de altura')
            self.height_cm = None

        try:
            self.weight_kg = float(input("Peso en Kg: "))
            while self.weight_kg <= 0:
                print('Ingrese un valor válido > 0')
                self.weight_kg = float(input("Peso en Kg: "))
        except ValueError:
            print('Error en ingreso de peso')
            self.weight_kg = None

        self.occupation = input("Ocupación: ")
        self.occupation = self.occupation.strip().title()

    def get_name(self):
        return self.name

    def get_age(self):
        return self.age

    def get_gender(self):
        return self.gender

    def get_weight_kg(self):
        return self.weight_kg

    def get_height_cm(self):
        return self.height_cm

    def get_occupation(self):
        return self.occupation


def create_directories():
    lam_dir = '~\\Documents\\LAM\\'
    patient_dir = lam_dir + patient_name + '\\'
    img_dir = patient_dir + 'Imagenes\\'

    lam_dir = path.expanduser(lam_dir)
    patient_dir = path.expanduser(patient_dir)
    img_dir = path.expanduser(img_dir)

    db_xlsx_dir = lam_dir + 'DB_LAM.xlsx'

    if not path.isdir(lam_dir):
        mkdir(lam_dir)

    if not path.isdir(patient_dir):
        mkdir(patient_dir)
        patient_dir = path.expanduser(patient_dir)

    if not path.isdir(img_dir):
        mkdir(img_dir)
        img_dir = path.expanduser(img_dir)

    return patient_dir, img_dir, db_xlsx_dir


def load_image(name_img):
    root = Tk()

    def load_button():

        global photo

        configuration = read_configuration()

        photo = filedialog.askopenfilename(title='Abrir', initialdir=configuration['last_dir'],
                                           filetypes=(('Imagenes', '*.jpg'),
                                                      ('Imagenes', '*.png'),
                                                      ('Todos los ficheros', '*.*')))

        configuration['last_dir'] = path.abspath(path.join(photo, '..'))
        with open('configuration.json', 'w') as file:
            dump(configuration, file, indent=4)

        root.destroy()

    try:
        root.iconbitmap('icon.ico')
    except TclError:
        print('\nNo ico file found')

    root.title('Análisis Postural LAM')

    root.geometry('300x50')
    Label(root, text=f'Cargue la image {name_img}').pack()
    Button(root, text='Cargar imagen', command=load_button).pack()

    root.mainloop()


def read_configuration():
    if path.isfile('configuration.json'):
        with open('configuration.json', 'r') as json_file:
            configuration = load(json_file)
    else:
        configuration = {"tolerance_angle": 0.0,
                         "tolerance_distance": 0.0,
                         "grid_size": 5.0,
                         "filter_threshold": 0.32,
                         "last_dir": getcwd()
                         }
    return configuration


def change_configuration():
    while True:
        system("cls")

        configuration = read_configuration()
        print('La configuración actual es:\n')

        i = 0
        parameter_names = ('Ángulo de tolerancia', 'Distancia de tolerance', 'Tamaño de cuadrícula',
                           'Valor umbral del filtro')
        for value in configuration.values():
            if i < 4:
                print(f'{i + 1}. {parameter_names[i]} = {value}')
                i += 1

        print(f'5. Restablecer configuración')
        print(f'6. Salir de configuracion')

        opt = input('\nQue parámetro desea configurar: ')

        if opt == '1':
            try:
                configuration['tolerance_angle'] = float(input('Ingrese un valor entre 0.0 y 5.0: '))
                while configuration['tolerance_angle'] <= -0.01 or configuration['tolerance_angle'] >= 5.01:
                    print('Valor no válido')
                    configuration['tolerance_angle'] = float(input('Ingrese un valor entre 0.0 y 5.0: '))
            except ValueError:
                print('Error de ingreso.')
                input()

        elif opt == '2':
            try:
                configuration['tolerance_distance'] = float(input('Ingrese un valor entre 0.0 y 3.0: '))
                while configuration['tolerance_distance'] <= -0.01 or configuration['tolerance_distance'] >= 3.01:
                    print('Valor no válido')
                    configuration['tolerance_distance'] = float(input('Ingrese un valor entre 0.0 y 3.0: '))
            except ValueError:
                print('Error de ingreso.')
                input()

        elif opt == '3':
            try:
                configuration['grid_size'] = float(input('Ingrese un valor entre 1.0 y 10.0: '))
                while configuration['grid_size'] <= -0.01 or configuration['grid_size'] >= 10.01:
                    print('Valor no válido')
                    configuration['grid_size'] = float(input('Ingrese un valor entre 1.0 y 10.0: '))
            except ValueError:
                print('Error de ingreso.')
                input()

        elif opt == '4':
            try:
                configuration['filter_threshold'] = float(input('Ingrese un valor entre 0.20 y 0.80: '))
                while configuration['filter_threshold'] <= 0.19 or configuration['filter_threshold'] >= 0.81:
                    print('Valor no válido')
                    configuration['filter_threshold'] = float(input('Ingrese un valor entre 0.20 y 0.80: '))
            except ValueError:
                print('Error de ingreso.')
                input()

        elif opt == '5':
            configuration['tolerance_angle'] = 0.0
            configuration['tolerance_distance'] = 0.0
            configuration['grid_size'] = 5.0
            configuration['filter_threshold'] = 0.32
            configuration['last_dir'] = getcwd()

        elif opt == '6':
            break

        else:
            print('El valor ingresado es incorrecto')
            input()

        with open('configuration.json', 'w') as file:
            dump(configuration, file, indent=4)


def menu():
    while True:
        system('cls')
        print(f'{"*" * 10} MENU {"*" * 10}\n')
        print('1. Hacer analisis postural')
        print('2. Abrir base de datos')
        print('3. Cambiar configuracion')
        print('4. Salir')

        option = input('\nIngrese una opcion: ')

        if option == '1':
            system('cls')

            configuration = read_configuration()

            tolerance_angle = configuration['tolerance_angle']
            tolerance_distance = configuration['tolerance_distance']
            grid_size = configuration['grid_size']
            filter_threshold = configuration['filter_threshold']

            postural_assessment(tolerance_angle, tolerance_distance, grid_size, filter_threshold)

        elif option == '2':
            _, _, db_xlsx_dir = create_directories()
            if path.isfile(db_xlsx_dir):
                open_new(db_xlsx_dir)
                input('Presione cualquier tecla para continuar... ')
            else:
                print('No se encontró el archivo')
                print(db_xlsx_dir)
                input('Presione cualquier tecla para continuar... ')

        elif option == '3':
            change_configuration()

        elif option == '4':
            break
        else:
            print('El valor ingresado es incorrecto')


# MAIN

if __name__ == '__main__':

    # Checking the usage time

    year = datetime.now().year
    month = datetime.now().month
    day = datetime.now().day

    current_date = date(year, month, day)
    due_date = date(2021, 6, 12)

    trial_days = (due_date - current_date).days

    if (due_date - current_date).days > 0:
        print(f'\nLe quedan {trial_days} días de prueba')
        input('Presione cualquier tecla para continuar... ')
    else:
        print(f'\nEl tiempo de prueba se ha terminado')

    if current_date < due_date:

        # globals variables
        patient_name = ''

        photo = ''

        run_anterior_assessment = False
        run_posterior_assessment = False
        run_lateral_r_assessment = False

        try:
            menu()
        except KeyboardInterrupt:
            print('\nCierre forzado')
        except ValueError:
            print('\nError inesperado')

    else:
        print(f'Finalizó el: {due_date}')
        input('Presione cualquier tecla para continuar... ')

# To create executable: pyinstaller --onefile --icon=icon.ico sistema_LAM.py
# Note: First the missing hook must be added (hook-skimage.filters.py) at the library \PyInstaller\hooks
