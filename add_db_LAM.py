# LIBRARIES
from time import strftime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pandas import DataFrame, ExcelWriter
from os import path


def transform_information_xlsx(img_dir, pdf_dir, patient_data, diagnostic, db_assessment):
    df1 = db_assessment['table_1']
    df2 = db_assessment['table_2']
    df3 = db_assessment['table_3']
    values_tables = df1.values.tolist() + df2.values.tolist() + df3.values.tolist()
    values_tables = [i[:][2] for i in values_tables]

    img_dir = [f'=HYPERLINK(\"{img_dir}\",\"{patient_data[1]}  {patient_data[0]}\")']

    data = img_dir + pdf_dir + patient_data + diagnostic + values_tables
    data = DataFrame(data)

    return data.T


def header_table_format(sheet_name, text, start_cell, end_cell):
    """
    :param sheet_name:
    :param text: Cell Value
    :param start_cell:
    :param end_cell:
    :return:
    """

    thick = Side(border_style='thick', color='000000')

    sheet_name.merge_cells(f'{start_cell}:{end_cell}')
    cell = sheet_name[start_cell]
    cell.value = text

    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True, italic=True)
    cell.fill = PatternFill('solid', fgColor='00C0C0C0')

    sheet_name.merge_cells(f'{start_cell}:{end_cell}')
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)


def create_xlsx(db_xlsx_dir):
    writer = ExcelWriter(db_xlsx_dir, engine='openpyxl')
    DataFrame().to_excel(writer, 'Anterior', index=False)
    DataFrame().T.to_excel(writer, 'Posterior', index=False)
    DataFrame().T.to_excel(writer, 'LateralD', index=False)
    writer.save()
    writer.close()

    wb = load_workbook(db_xlsx_dir)
    wb.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    name_sheets = list(writer.sheets.keys())

    # ANTERIOR

    sheet = wb[name_sheets[0]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'O1')
    header_table_format(sheet, 'Grados con la horizontal', 'P1', 'R1')
    header_table_format(sheet, 'Distancia con la vertical', 'S1', 'X1')
    header_table_format(sheet, 'Rotacion de los pies', 'Y1', 'Z1')

    # POSTERIOR

    sheet = wb[name_sheets[1]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'M1')
    header_table_format(sheet, 'Grados con la horizontal', 'N1', 'P1')
    header_table_format(sheet, 'Distancia con la vertical', 'Q1', 'V1')
    header_table_format(sheet, 'Rotacion de los pies', 'W1', 'X1')

    # LATERAL R

    sheet = wb[name_sheets[2]]

    header_table_format(sheet, 'Hepervínculos', 'A1', 'B1')
    header_table_format(sheet, 'Datos personales', 'C1', 'I1')
    header_table_format(sheet, 'Diagnóstico', 'J1', 'K1')
    header_table_format(sheet, 'Grados con la vertical', 'L1', 'O1')
    header_table_format(sheet, 'Grados con la horizontal', 'P1', 'P1')
    header_table_format(sheet, 'Distancia con la vertical', 'Q1', 'U1')

    wb.save(db_xlsx_dir)
    wb.close()

    with ExcelWriter(db_xlsx_dir, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        anterior_header = DataFrame((), ('Imagen', 'Reporte',

                                         'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                         'Linea escapular descendida', 'Linea pélvica descendida',
                                         'Cabeza inclinada', 'Cabeza rotada', 'Rodilla derecha', 'Rodilla izquierda',

                                         'Hombro descendido', 'Pelvis descendida', 'Rodilla descendida',

                                         'Frente', 'Hombros', 'Ombligo', 'Pelvis', 'Rodillas', 'Pies',

                                         'Pie izquierdo', 'Pie derecho'))

        posterior_header = DataFrame((), ('Imagen', 'Reporte',

                                          'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                          'Rotación dorsal posterior', 'Rotación pélvica posterior', 'Pie izquierdo',
                                          'Pie derecho',

                                          'Hombro descendido', 'Pelvis descendida', 'Rodilla descendida',

                                          'Hombros', '7ma cervical', '5ta torácica', 'Pelvis', 'Rodillas',
                                          'Tobillos',

                                          'Pie izquierdo', 'Pie derecho',))

        lateral_r_header = DataFrame((), ('Imagen', 'Rreporte',

                                          'Fecha', 'Nombre', 'Edad', 'Género', 'Talla[cm]', 'Peso[kg]', 'Ocupación',

                                          'Postura', 'Proyección postural',

                                          'Cabeza-hombro', 'Hombro-pelvis', 'Cadera-rodilla', 'Rodilla-pie',

                                          'Pelvis',

                                          'Cabeza', 'Hombro', 'Pelvis', 'Cadera', 'Rodilla'))

        anterior_header.T.to_excel(writer, name_sheets[0], header=True, index=False, startrow=1, startcol=0)
        posterior_header.T.to_excel(writer, name_sheets[1], header=True, index=False, startrow=1, startcol=0)
        lateral_r_header.T.to_excel(writer, name_sheets[2], header=True, index=False, startrow=1, startcol=0)


def add_data_db_xlsx(db_xlsx_dir, patient, data_dict, img_dict, report_pdf_dir, books):
    if not path.isfile(db_xlsx_dir):
        create_xlsx(db_xlsx_dir)

    assessment_date = strftime('%d/%m/%Y')
    db_anterior = data_dict['Anterior']
    db_posterior = data_dict['Posterior']
    db_lateral_r = data_dict['Lateral_d']

    anterior_img_dir = img_dict['Anterior']
    posterior_img_dir = img_dict['Posterior']
    lateral_r_img_dir = img_dict['Lateral_d']

    patient_name = patient.get_name()
    patient_age = patient.get_age()
    patient_gender = patient.get_gender()
    weight_kg = patient.get_weight_kg()
    height_cm = patient.get_height_cm()
    occupation = patient.get_occupation()

    anterior_sheet = books[0]
    posterior_sheet = books[1]
    lateral_r_sheet = books[2]

    patient_data = [assessment_date, patient_name, patient_age, patient_gender, height_cm, weight_kg, occupation]
    report_pdf_dir = [f'=HYPERLINK(\"{report_pdf_dir}\",\"{patient_name}  {assessment_date}\")']
    wb = load_workbook(db_xlsx_dir)

    try:

        with ExcelWriter(db_xlsx_dir, engine='openpyxl') as writer:

            writer.book = wb
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            name_sheets = list(writer.sheets.keys())

            if anterior_sheet:
                sheet = wb[name_sheets[0]]
                max_row = sheet.max_row

                dg = db_anterior['diagnostic']
                dg = [dg[0], dg[1], dg[3], dg[4], dg[6], dg[8]]

                data = transform_information_xlsx(anterior_img_dir, report_pdf_dir, patient_data, dg, db_anterior)
                data.to_excel(writer, name_sheets[0], header=False, index=False, startrow=max_row, startcol=0)

            if posterior_sheet:
                sheet = wb[name_sheets[1]]
                max_row = sheet.max_row

                dg = db_posterior['diagnostic']
                dg = [dg[0], dg[1], dg[2], dg[3]]

                data = transform_information_xlsx(posterior_img_dir, report_pdf_dir, patient_data, dg, db_posterior)
                data.to_excel(writer, name_sheets[1], header=False, index=False, startrow=max_row, startcol=0)

            if lateral_r_sheet:
                sheet = wb[name_sheets[2]]
                max_row = sheet.max_row

                dg = db_lateral_r['diagnostic']
                dg = [dg[0], dg[1]]

                data = transform_information_xlsx(lateral_r_img_dir, report_pdf_dir, patient_data, dg, db_lateral_r)
                data.to_excel(writer, name_sheets[2], header=False, index=False, startrow=max_row, startcol=0)

        wb.close()
        print('\nDatos agregados con éxito')

    except PermissionError:
        wb.close()
        print('\nError de escritura')
        print('Revise que el archivo no se encuentre abierto o los permisos del antivirus\n')
        reload_data = bool(int(input("¿Desea cargar nuevamente? 1/0: ")))

        if reload_data:
            add_data_db_xlsx(db_xlsx_dir, patient, data_dict, img_dict, report_pdf_dir, books)
